import openpyxl
from pathlib import Path
p = Path('c:/Users/love6/OneDrive/문서/GitHub/overlaylist/LIST.xlsx')
if not p.exists():
    raise FileNotFoundError(p)
wb = openpyxl.load_workbook(p, data_only=True)
print('sheets=', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('--- SHEET:', name)
    rows = list(ws.iter_rows(values_only=True))
    print('rowcount=', len(rows))
    for i, row in enumerate(rows[:20], 1):
        print(i, row)
